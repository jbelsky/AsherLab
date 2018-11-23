# -*- coding: utf-8 -*-
"""
Created on Thu Nov 15 17:07:00 2018

@author: BelskyJ
"""

import openpyxl
import pandas as pd
import numpy as np
import os
import glob
import argparse
import re
from collections import OrderedDict

import classmatrix

# Set up the new OPClass

class OParkClass:

	def __init__(self, _classID, _students_list, _friendship_mat, _allClassPeerProvisionsByItem_dict):

		self.classID = _classID
		self.students_list = _students_list
		self.friendships_mat = _friendship_mat

		self.peerProvisions_dict = OrderedDict()
		self.peerProvisions_items = list(_allClassPeerProvisionsByItem_dict)
		self.peerProvisions_items.sort()

		for i in items:
			if self.classID in _allClassPeerProvisionsByItem_dict[i]:
				self.peerProvisions_dict[i] = _allClassPeerProvisionsByItem_dict[i][self.classID]

	def get_studentIDs(self):

		return [s.studentID for s in self.students_list]

	def get_studentID_index(self, _studentID):

		studentID_list = self.get_studentIDs()
		try:
			return studentID_list.index(_studentID)
		except:
			print("StudentID %d is not in %s" % (_studentID, self.classID))

	def get_peer_provisions_items(self):

		return self.peerProvisions_items

	def get_peer_provisions_item_df(self, _peerProvItem):

		return self.peerProvisions_dict[_peerProvItem]

	def get_friendship_mat(self):
		return self.friendships_mat

# Set up the new class
class OParkStudent:

	def __init__(self, _studentID, _gender):

		self.studentID = _studentID
		self.gender = _gender

	def set_received_peerprov_mat(self, _OPclass):

		self.receivedPeerProv_mat = pd.DataFrame(index = _OPclass.get_studentIDs(),
											   columns = list(_OPclass.peerProvisions_dict.keys()),
											   dtype = int
											  )

		# Fill in the data frame from the OParkClass
		for item, prov_df in _OPclass.peerProvisions_dict.items():
			self.receivedPeerProv_mat.loc[:,item] = prov_df.loc[:,self.studentID]

	def set_received_friendships(self, _OPclass):

		friendship_mat = _OPclass.get_friendship_mat()
		i = _OPclass.get_studentID_index(self.studentID)
		self.receivedFriendships = friendship_mat[i,:]



def GetPeerProvisions(_dir):

	# Initialize the peerProvision matrix
	classPeerProvisionByItem_dict = {}

	# Obtain the number of peer provision matrices
	peerProv_xlsx_files = glob.glob(_dir + "/*xlsx")
	peerProv_xlsx_files.sort()

	# Get the peer provision item -> file dict
	p = re.compile("Peer provisions item (\d{1,2})_.*\.xlsx")

	for f in peerProv_xlsx_files:

		print(f)

		item = int(p.match(os.path.basename(f)).group(1))
		classPeerProvisionByItem_dict[item] = {}

		pp_wb = openpyxl.load_workbook(f)

		for OPclass in pp_wb.sheetnames:

			if "Class" not in OPclass:
				continue

			data_df, gender_s = classmatrix.GetDataMatrix(pp_wb[OPclass])

			# Enter into dict
			classPeerProvisionByItem_dict[item][OPclass] = data_df

		pp_wb.close()

	return classPeerProvisionByItem_dict


def GetClassFriendshipMatrix(_friendship_df):

	# Initialize the class_friendship_matrix
	classFriendship_mat = np.empty(_friendship_df.shape, dtype = object)

	# Iterate through each row
	for i in range(0, classFriendship_mat.shape[0]):

		for j in range(classFriendship_mat.shape[1]):

			# Get the given and received status
			isGiven = True if _friendship_df.iloc[i, j] == 1 else False
			isReceived = True if _friendship_df.iloc[i, j] == 1 else False

			if isGiven and isReceived:
				classFriendship_mat[i, j] = "reciprocated"
			elif isGiven:
				classFriendship_mat[i, j] = "given"
			elif isReceived:
				classFriendship_mat[i, j] = "received"
			else:
				classFriendship_mat[i, j] = "none"

	# Return the friendship matrix
	return classFriendship_mat

parser = argparse.ArgumentParser()
parser.add_argument("friendship_nom_file", help = "Friendship Nominations File (.xlsx)")
parser.add_argument("peer_provisions_dir", help = "Directory containing peer provisions")
parser.add_argument("-o", "--output", action = "store", default = "compare_excel_dirs.txt", type = str)
args = parser.parse_args()

# Get the peer provisions for each of the classes
classPeerProvisionsByItem_dict = GetPeerProvisions(args.peer_provisions_dir)

# Load in the friendship matrix
wb = openpyxl.load_workbook(args.friendship_nom_file)
class_sn = wb.sheetnames
class_sn.sort()

# Set up the friendship_matrix dict
OPclass_dict = OrderedDict()

for c in class_sn:

	# Import the nominations matrix
	friendship_df, gender_srs = classmatrix.GetDataMatrix(wb[c])

	# Get the friendship matrix
	friendship_mat = GetClassFriendshipMatrix(friendship_df)

	# Set the OParkStudent list
	OPstudent_list = [OParkStudent(i, gender_srs.loc[i]) for i in gender_srs.index]

	# Initialize an OPClass
	orlandParkClass = OParkClass(c, OPstudent_list, friendship_mat, classPeerProvisionsByItem_dict)

	break

wb.close()


ops = OParkStudent(1502, 1)
ops.set_received_peerprov_mat(orlandParkClass)
ops.set_received_friendships(orlandParkClass)


'''
# Initialize the provisions_matrix dict
provisionsByClass_dict = OrderedDict()



# Iterate through each peer provision
for pp_xls in peerProv_xlsx_files:

	pp_wb = openpyxl.load_workbook(pp_xls)

	# Iterate through each class_sn
	for c in class_sn:

		try:
			data_df, gender_s = classmatrix.GetDataMatrix(pp_wb[c])
		except:
			print("%s is not in %s, skipping..." % (c, pp_xls))
			continue


# Load in the peer provisions
for pp_xls in peerProv_xls_files:

	pp_wb = openpyxl.load_workbook(pp_xls)

	for sheet_name in pp_wb.sheetnames:

		if "Class" not in sheet_name:
			continue

		print("%s\t%s" % (os.path.basename(pp_xls), sheet_name))



		break

	break

'''